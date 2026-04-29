from pathlib import Path
import pdfplumber
from docx import Document


def read_file(path: str) -> str:
    """
    Extract text from a downloaded tender file.
    Supported: PDF, DOCX, XLSX, TXT.
    Returns empty string on unsupported type or failure.
    """
    ext = Path(path).suffix.lower()
    try:
        if ext == ".pdf":
            return _read_pdf(path)
        elif ext in (".docx", ".doc"):
            return _read_docx(path)
        elif ext in (".xlsx", ".xls"):
            return _read_excel(path)
        elif ext == ".txt":
            with open(path, encoding="utf-8", errors="ignore") as f:
                return f.read()
        else:
            return ""
    except Exception as e:
        return f"[Could not read {Path(path).name}: {e}]"


def _read_pdf(path: str) -> str:
    parts = []
    with pdfplumber.open(path) as pdf:
        for pg in pdf.pages:
            # Regular text
            text = pg.extract_text()
            if text:
                parts.append(text)
            # Tables — eligibility criteria and budget are often in tables
            try:
                for table in pg.extract_tables():
                    for row in table:
                        row_str = " | ".join(str(c or "").strip() for c in row if c)
                        if row_str.strip():
                            parts.append(f"[TABLE] {row_str}")
            except Exception:
                pass
    return "\n".join(parts)


def _read_docx(path: str) -> str:
    doc = Document(path)
    parts = []

    from docx.text.paragraph import Paragraph
    from docx.table import Table

    # Walk the document body in order: paragraphs and tables are interleaved
    for block in doc.element.body:
        tag = block.tag.split("}")[-1] if "}" in block.tag else block.tag
        if tag == "p":
            para = Paragraph(block, doc)
            if para.text.strip():
                parts.append(para.text.strip())
        elif tag == "tbl":
            tbl = Table(block, doc)
            for row in tbl.rows:
                row_text = " | ".join(cell.text.strip() for cell in row.cells if cell.text.strip())
                if row_text:
                    parts.append(f"[TABLE] {row_text}")

    return "\n".join(parts)


def _read_excel(path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        parts.append(f"[Sheet: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            row_text = "  |  ".join(str(c) for c in row if c is not None)
            if row_text.strip():
                parts.append(row_text)
    wb.close()
    return "\n".join(parts)
