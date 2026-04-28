from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _c(ws, row, col, value=None, bold=False, size=11, h_align=None, v_align="top", wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=size)
    cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
    cell.border = _BORDER
    return cell


def _write_level1_sheet(ws, record: dict, sourced_on: str):
    f = record.get("fields", {})
    title = record.get("title", "")
    site = record.get("site", "")
    url = record.get("url", "")

    def r(row_num, a=None, b=None, c=None, d=None, a_bold=False, c_bold=False):
        _c(ws, row_num, 1, a, bold=a_bold, v_align="top")
        _c(ws, row_num, 2, b, h_align="left", v_align="top")
        _c(ws, row_num, 3, c, bold=c_bold, h_align="center", v_align="top")
        _c(ws, row_num, 4, d, v_align="top")

    # Row 1 — column headers
    _c(ws, 1, 1, "Category", bold=True, size=12, v_align="top")
    _c(ws, 1, 2, "Data",     bold=True, size=12, v_align="top")
    _c(ws, 1, 3, None,       bold=True, size=12)
    _c(ws, 1, 4, "Remarks",  bold=True, size=11, v_align="top")
    ws.row_dimensions[1].height = 22.5

    # ── Level 1 section ──────────────────────────────────────────────────────
    r(2,  "Level 1",        "Source",                    site,                              url,                              a_bold=True)
    r(3,  "Identification", "Published on",              f.get("PUBLISHED_ON", ""),         None,                             a_bold=True, c_bold=True)
    r(4,  None,             "Sourced on",                sourced_on,                        None,                             c_bold=True)
    r(5,  None,             "Pre-bid Meeting",           f.get("PRE_BID_MEETING", "Not specified"))
    r(6,  None,             "Last date for Submission",  f.get("DEADLINE", ""),             None,                             c_bold=True)
    # D7 = Reference No. (Remarks for Bid Title row)
    r(7,  None,             "Bid Title",                 f.get("BID_TITLE") or title,       f.get("REFERENCE_NO", ""))
    # D8 = Contact Email (Remarks for Inviting Authority row)
    r(8,  None,             "Inviting Authority",        f.get("INVITING_AUTHORITY", ""),   f.get("CONTACT_EMAIL", ""),       c_bold=True)
    r(9,  None,             "Funding Agency",            f.get("FUNDING_AGENCY", ""),       None,                             c_bold=True)
    r(10, None,             "Country of Assignment",     f.get("COUNTRY", ""))
    r(11, None,             "Domain",                    f.get("DOMAIN", ""))
    r(12, None,             "Sub Domain",                f.get("SUB_DOMAIN", ""))
    r(13, None,             "TMI Service Line",          f.get("TMI_SERVICE_LINE", ""))
    r(14, None,             "Estimated Project Value",   f.get("PROJECT_VALUE", ""))
    # D15 = full eligibility details
    r(15, None,             "Eligibility",               f.get("ELIGIBILITY_FLAG", ""),     f.get("ELIGIBILITY_DETAILS", ""))
    # D16 = dependencies + access/platform requirements combined
    _dep = f.get("DEPENDENCIES", "")
    _acc = f.get("ACCESS_REQUIREMENTS", "")
    _dep_combined = " | ".join(x for x in [_dep, _acc] if x and x.lower() != "not specified") or ""
    r(16, None,             "Dependencies",              None,                              _dep_combined)
    r(17, None,             "Consortium",                f.get("CONSORTIUM", ""))
    # Row 18 — Scope of Work (wrap + taller)
    _c(ws, 18, 1, None, v_align="top")
    _c(ws, 18, 2, "Scope of Work", h_align="left", v_align="top")
    _c(ws, 18, 3, f.get("SCOPE_OF_WORK", ""), v_align="top", wrap=True)
    _c(ws, 18, 4, None, v_align="top")
    ws.row_dimensions[18].height = 45
    r(19, None, "Delivery Period",               f.get("DELIVERY_PERIOD", ""))
    r(20, None, "EMDs / Deposits if any",        f.get("EMDS", ""))
    # D21 = pricing conditions (Remarks for Payment Terms row)
    r(21, None, "Payment Terms",                 f.get("PAYMENT_TERMS", ""),              f.get("PRICING_CONDITIONS", ""))
    r(22, None, "DU",                            None)   # user fills
    r(23, None, "Selection Criteria (Proposal)", f.get("SELECTION_CRITERIA", ""))

    # ── Level 2 section ──────────────────────────────────────────────────────
    r(24, "Level 2",    "Clarity on the Scope",        f.get("SCOPE_CLARITY", ""),   None, a_bold=True)
    r(25, "Validation", "Scope - Quantum & related",   f.get("SCOPE_QUANTUM", ""),   None, a_bold=True)
    r(26, None,         "Earlier experience in TMI",   None,                         None)   # user fills
    r(27, None,         "Clarifications if any",       None,                         f.get("CLARIFICATIONS", ""))
    r(28, None,         "EMDs/ Deposits if any",       f.get("EMDS", ""))
    r(29, None,         "Impact of Payment Terms if any", None)
    r(30, None,         "Penalities & LDs",            f.get("PENALTIES", ""))

    # ── Level 3 section ──────────────────────────────────────────────────────
    r(31, "Level 3",        "Decided By",               None, a_bold=True)
    r(32, "Decision",       "Date & Time",              None, a_bold=True)
    r(33, "BU / DU / SPOC", "Last date for Submission", f"Submission: {f.get('DEADLINE', '')}", a_bold=True)
    r(34, None,             "Action Plan",              None)
    for i, n in enumerate(range(1, 6), 35):
        r(i, None, n, None)

    # Column widths matching the Level 1 template
    ws.column_dimensions["A"].width = 22.5
    ws.column_dimensions["B"].width = 29.0
    ws.column_dimensions["C"].width = 65.0
    ws.column_dimensions["D"].width = 68.0


def _safe_sheet_name(title: str, existing: list) -> str:
    """Return a valid, unique Excel sheet name (max 31 chars)."""
    name = title[:28].strip()
    for ch in r'\/:*?[]':
        name = name.replace(ch, " ")
    name = name.strip() or "Tender"
    base, suffix = name, 1
    while name in existing:
        name = f"{base[:25]}_{suffix}"
        suffix += 1
    return name


def write_level1_report(records: list, output_path: str) -> None:
    """Create one Excel file with one Level-1 sheet per tender record."""
    wb = Workbook()
    wb.remove(wb.active)   # discard the default blank sheet

    sourced_on = datetime.now().strftime("%d-%b-%Y")
    used_names: list[str] = []

    for record in records:
        sheet_name = _safe_sheet_name(record.get("title", "Tender"), used_names)
        used_names.append(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        _write_level1_sheet(ws, record, sourced_on)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
