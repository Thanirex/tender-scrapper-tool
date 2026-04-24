import json
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from dotenv import load_dotenv

# Central Agents
from agents.scraper_agent import ScraperAgent
from agents.summarizer_agent import SummarizerAgent

load_dotenv()

def format_excel(excel_path):
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    
    # Header format
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    align_center = Alignment(horizontal="center", vertical="center")
    align_wrap = Alignment(wrap_text=True, vertical="top")

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    # Column Widths
    sheet.column_dimensions["A"].width = 6   # S.no
    sheet.column_dimensions["B"].width = 15  # Website
    sheet.column_dimensions["C"].width = 20  # Keyword
    sheet.column_dimensions["D"].width = 40  # Document/Title Name
    sheet.column_dimensions["E"].width = 80  # Summary
    
    # Wrapping text on summary
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=5).alignment = align_wrap

    workbook.save(excel_path)

def main():
    print("🚀 Starting V2 Agentic Scraper...")
    
    # 1. Load keywords from JSON
    try:
        with open("Keywords.json", "r") as f:
            keyword_data = json.load(f)
    except FileNotFoundError:
        print("❌ Keywords.json not found!")
        return

    # Load ALL keywords from both C&K and HR&KM categories
    ck_keywords = keyword_data.get("C&K", [])
    hrkm_keywords = keyword_data.get("HR&KM", [])
    test_keywords = ck_keywords + hrkm_keywords

    print(f"✅ Loaded {len(ck_keywords)} C&K + {len(hrkm_keywords)} HR&KM = {len(test_keywords)} total keywords.")

    # 2. Setup Database
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"data/outputs/Scrape_Results_{timestamp}.xlsx"
    columns = ["S.no", "Website", "Keyword", "Document Name", "Summary"]
    
    # Keep track of records
    records = []
    s_no = 1
    
    # 3. The actual scraper loop
    agent = ScraperAgent("sites_config.json")
    summarizer = SummarizerAgent()
    
    for keyword in test_keywords:
        results = agent.search("devnet", keyword)
        if not results:
            print(f"   ↳ No results found for '{keyword}'")
            continue
            
        for res in results:
            summary = summarizer.summarize(res.get('content', ''))
            records.append([s_no, "devnet", keyword, res.get('title', 'Unknown'), summary])
            s_no += 1

    # 4. Save and Format using pure openpyxl (No Pandas required!)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    for row in records:
        ws.append(row)
    wb.save(output_path)
    
    format_excel(output_path)
    print(f"✅ Scrape completed and formatted at {output_path}")

if __name__ == "__main__":
    main()
